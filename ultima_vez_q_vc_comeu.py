import openpyxl
from datetime import datetime
from openpyxl.styles import Alignment
# Propósito do script:
# Ando comendo muito, muita porcaria, então quero espaçar minhas refeições.
# Para isso, criei esse script,
# que é basicamente um "botão" de registro da última refeição feita
# No caso, o registro vai para um arquivo de Excel.





# verificação da existencia do arquivo. se n existe, cria, se existe, abre
try:
    workbook = openpyxl.load_workbook("tabela_horários.xlsx")
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = "Última vez que você comeu"
    sheet.column_dimensions["A"].width = 20
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")

# verificação de qual a ultima linha preenchida
last_row = sheet.max_row

# bloco que, de fato, adiciona o horário a cada enter pressionado
row = last_row + 1
while True:
    input("Pressione Enter para adicionar o horário atual à tabela:")
    sheet.cell(row=row, column=1, value=datetime.now().strftime("%d-%m-%Y %H:%M"))
    sheet.cell(row=row, column=1).alignment = Alignment(horizontal="center")

    row += 1
    
    # salva o arquivo
    workbook.save("tabela_horários.xlsx")
    
    


